from __future__ import annotations

import concurrent.futures
import logging
import os
import re
import sys
import tempfile
import threading
from collections.abc import Callable
from typing import TYPE_CHECKING

import pandas as pd
from patchright.sync_api import sync_playwright

from scraper.dedup import deduplicate
from scraper.pipeline import extract_company_info
from scraper.search import google_search, load_search_cache, save_search_cache
from scraper.types import SearchResult

if TYPE_CHECKING:
    from scraper.config import Config

logger = logging.getLogger(__name__)

NUM_WORKERS = 4
COLUMNS = [
    "Company Name",
    "Contact Person",
    "Position",
    "State",
    "City",
    "Country",
    "Phone Number",
    "Email",
    "Website",
    "Products",
]


def _check_cancelled(cancel_event: threading.Event | None) -> None:
    if cancel_event is not None and cancel_event.is_set():
        raise KeyboardInterrupt("Cancelled by user")


def _save_checkpoint(df: pd.DataFrame, output_path: str) -> None:
    try:
        import openpyxl  # noqa: F401
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter

        thin_border = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9"),
        )
        header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        not_found_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

        styled = df.style.set_table_styles(
            [
                {"selector": "th", "props": [("font-weight", "bold"), ("text-align", "left")]},
                {"selector": "td", "props": [("text-align", "left")]},
            ],
            overwrite=False,
        )

        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            styled.to_excel(writer, index=False, sheet_name="Suppliers")
            ws = writer.sheets["Suppliers"]

            for col_idx, col in enumerate(ws.columns, 1):
                max_len = 0
                for cell in col:
                    try:
                        val_len = len(str(cell.value or ""))
                        if val_len > max_len:
                            max_len = val_len
                    except Exception:
                        pass
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 60)

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.border = thin_border

            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical="center")
                    if cell.value and isinstance(cell.value, str) and cell.value.strip() == "Not Found":
                        cell.fill = not_found_fill

            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions

        logger.info("Styled Excel saved: %s", output_path)
    except ImportError:
        csv_path = output_path.rsplit(".", 1)[0] + ".csv"
        df.to_csv(csv_path, index=False)
        logger.warning("openpyxl not available, saved CSV: %s", csv_path)


def _domain_fallback_name(url: str) -> str:
    from urllib.parse import urlparse

    parsed = urlparse(url)
    domain = parsed.netloc or parsed.path or ""
    domain = domain.lower().removeprefix("www.").removesuffix("/")
    parts = [p for p in domain.split(".") if p]
    skip = {"com", "co", "org", "net", "gov", "edu", "uk", "za", "in", "au"}
    for p in reversed(parts):
        if p not in skip and len(p) >= 3:
            return p.capitalize()
    return parts[0].capitalize() if parts else url


def _is_valid_url(url: str) -> bool:
    if not url.startswith(("http://", "https://")):
        return False
    if "google.com/search" in url or url.startswith("/search"):
        return False
    if re.search(r"(?:^|://)(?:www\.)?(?:facebook|twitter|x)\.com/", url):
        return False
    if url.endswith(".pdf"):
        return False
    return True


def _process_chunk(
    company_items: list[tuple[str, SearchResult]],
    config: Config,
    worker_id: int,
) -> list[dict[str, str]]:
    pw = sync_playwright().start()
    browser = pw.chromium.launch(channel="chrome", headless=False)
    chunk_results: list[dict[str, str]] = []

    try:
        context = browser.new_context()
        page = context.new_page()
        page.set_default_timeout(config.page_load_timeout * 1000)
        page.route("**/*.{png,jpg,jpeg,gif,svg,ico,webp,woff,woff2,ttf,eot}", lambda route: route.abort())

        for company_name, search_result in company_items:
            if not _is_valid_url(search_result.url):
                logger.info("Skipping invalid URL: %s", search_result.url)
                continue
            try:
                contact = extract_company_info(page, search_result.url, company_name, config)
                if contact is not None:
                    chunk_results.append(
                        {
                            "Company Name": contact.company_name
                            if contact.company_name != "Not Found"
                            else _domain_fallback_name(search_result.url),
                            "Contact Person": contact.contact_person,
                            "Position": contact.position,
                            "State": contact.state,
                            "City": contact.city,
                            "Country": contact.country,
                            "Phone Number": contact.phone,
                            "Email": contact.email,
                            "Website": search_result.url,
                            "Products": contact.products,
                        }
                    )
            except Exception:
                logger.exception("Worker %d failed on: %s", worker_id, company_name)
        try:
            context.close()
        except Exception:
            pass
    finally:
        try:
            browser.close()
        except Exception:
            pass
        try:
            pw.stop()
        except Exception:
            pass

    return chunk_results


def run_scraper(
    config: Config,
    on_progress: Callable | None = None,
    cancel_event: threading.Event | None = None,
) -> pd.DataFrame:
    _setup_logging(config)
    original_cwd = os.getcwd()
    temp_dir = tempfile.mkdtemp(prefix="scraper_")
    os.chdir(temp_dir)

    all_rows: list[dict[str, str]] = []

    try:
        _check_cancelled(cancel_event)

        # Phase 1: Search
        cache_dir = config.cache_dir
        config_hash = config.search_config_hash()
        all_results = load_search_cache(cache_dir, config_hash, config.search_queries)

        if all_results is not None:
            logger.info("Loaded %d companies from search cache (%s)", len(all_results), cache_dir)
        else:
            logger.info("No valid cache — running Google search")
            all_results = {}

            pw = sync_playwright().start()
            browser = pw.chromium.launch(channel="chrome", headless=False)
            page = browser.new_page()
            page.set_default_timeout(config.page_load_timeout * 1000)
            page.route("**/*.{png,jpg,jpeg,gif,svg,ico,webp,woff,woff2,ttf,eot}", lambda route: route.abort())
            try:
                for query in config.search_queries:
                    logger.info("=== Search query: %s ===", query)
                    query_results = google_search(page, query, config)
                    save_search_cache(cache_dir, query, query_results, config_hash)
                    for name, result in query_results.items():
                        if name not in all_results:
                            all_results[name] = result
            finally:
                try:
                    browser.close()
                except Exception:
                    pass
                try:
                    pw.stop()
                except Exception:
                    pass

        logger.info("Found %d unique companies", len(all_results))
        _check_cancelled(cancel_event)

        items = list(all_results.items())
        if not items:
            return pd.DataFrame(columns=COLUMNS)

        _GENERIC_TITLES = {
            "about",
            "about us",
            "products",
            "product",
            "packaging",
            "home",
            "home page",
            "contact",
            "contact us",
            "services",
            "our services",
            "overview",
            "company name",
            "our products",
            "all products",
            "categories",
            "shop",
            "store",
            "our company",
        }
        items = [(n, r) for n, r in items if n.strip().lower() not in _GENERIC_TITLES]
        if not items:
            return pd.DataFrame(columns=COLUMNS)

        # Phase 2: Parallel company visits
        _check_cancelled(cancel_event)
        num_items = len(items)
        num_workers = min(NUM_WORKERS, num_items)
        chunks: list[list] = [[] for _ in range(num_workers)]
        for i, item in enumerate(items):
            chunks[i % num_workers].append(item)

        with concurrent.futures.ThreadPoolExecutor(max_workers=num_workers) as executor:
            futures = {
                executor.submit(_process_chunk, chunk, config, wid): wid
                for wid, chunk in enumerate(chunks) if chunk
            }
            completed_chunks = 0
            for future in concurrent.futures.as_completed(futures):
                if cancel_event is not None and cancel_event.is_set():
                    logger.info("Cancellation requested — shutting down workers")
                    executor.shutdown(wait=False, cancel_futures=True)
                    break

                try:
                    chunk_rows = future.result()
                    all_rows.extend(chunk_rows)
                except KeyboardInterrupt:
                    raise
                except Exception:
                    logger.exception("Worker chunk failed")

                completed_chunks += 1
                if on_progress:
                    temp_df = pd.DataFrame(all_rows, columns=COLUMNS)
                    on_progress(
                        temp_df,
                        len(all_rows),
                        num_items,
                        True,
                        f"chunk {completed_chunks}/{len(futures)}",
                    )

        _check_cancelled(cancel_event)
        df = pd.DataFrame(all_rows, columns=COLUMNS)

    except KeyboardInterrupt:
        logger.info("Scraper cancelled by user")
        df = pd.DataFrame(all_rows, columns=COLUMNS) if all_rows else pd.DataFrame(columns=COLUMNS)
        _save_checkpoint(df, os.path.join(original_cwd, config.output_filename))
        return df
    finally:
        os.chdir(original_cwd)
        import shutil

        try:
            shutil.rmtree(temp_dir)
        except Exception:
            pass

    df = deduplicate(df, name_col="Company Name", url_col="Website")
    output_path = os.path.join(original_cwd, config.output_filename)
    _save_checkpoint(df, output_path)
    logger.info("Done — %d companies saved to %s", len(df), output_path)
    return df


def _setup_logging(config: Config) -> None:
    root = logging.getLogger()
    root.setLevel(logging.INFO)
    if not root.handlers:
        handler = logging.StreamHandler(sys.stdout)
        handler.setLevel(logging.INFO)
        handler.setFormatter(
            logging.Formatter(
                "%(asctime)s [%(levelname)s] %(name)s: %(message)s",
                datefmt="%H:%M:%S",
            )
        )
        root.addHandler(handler)

        cache_dir = config.cache_dir
        os.makedirs(cache_dir, exist_ok=True)
        log_path = os.path.join(cache_dir, config.log_filename)
        file_handler = logging.FileHandler(log_path, mode="a", encoding="utf-8")
        file_handler.setLevel(logging.INFO)
        file_handler.setFormatter(
            logging.Formatter(
                "%(asctime)s [%(levelname)s] %(name)s: %(message)s",
                datefmt="%H:%M:%S",
            )
        )
        root.addHandler(file_handler)
        logger.info("Logging to %s", log_path)

    logger.info("Scraper started — queries: %d, workers: %d", len(config.search_queries), NUM_WORKERS)
